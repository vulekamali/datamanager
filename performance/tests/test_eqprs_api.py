from django.urls import reverse
from rest_framework.test import APITestCase, APIRequestFactory
from rest_framework import status
from django.test import Client
from openpyxl import load_workbook

import io


class indicator_API_Test(APITestCase):
    list_url = "/performance/api/v1/eqprs/"
    fixtures = ["test_api.json"]

    def test_api(self):
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, 200)
        response_payload = self.client.get(self.list_url).json()

        found = False
        message1 = "1.2.1 Percentage of valid invoices paid within 30 days upon receipt by the department"
        message2 = "1.1.1 Unqualified audit opinion"
        if (
            response_payload["results"]["items"][0]["indicator_name"] == message2
            or response_payload["results"]["items"][1]["indicator_name"] == message1
        ):
            found = True

        self.assertEqual(found, True)
        self.assertTrue(found, True)

    def test_create(self):
        response = self.client.post(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)

    def test_update(self):
        response = self.client.patch(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)

    def test_text_search(self):
        filter_url = self.list_url + "?page=1&q=Unqualified%20audit%20opinion"
        response_payload = self.client.get(filter_url).json()
        self.assertEqual(len(response_payload["results"]["items"]), 1)

    def test_frequency_search(self):
        filter_url = self.list_url + "?page=1&frequency=Annually"
        response_payload = self.client.get(filter_url).json()
        self.assertEqual(len(response_payload["results"]["items"]), 1)

    def test_downloaded_file(self):
        file_url = "/performance/performance-indicators.xlsx/"
        response = self.client.get(file_url, stream=True)
        file_content = response.getvalue()
        file_obj = io.BytesIO(file_content)
        wb = load_workbook(file_obj)
        sh = wb["Sheet1"]

        self.assertEqual(sh["A1"].value, "Financial year")
        self.assertEqual(sh["A2"].value, "2015-16")
        self.assertEqual(sh["H3"].value, "Annually")
        self.assertEqual(sh["A4"].value, None)

    def test_downloaded_file_with_filter(self):
        file_url = "/performance/performance-indicators.xlsx/?frequency=Annually"
        response = self.client.get(file_url, stream=True)
        file_content = response.getvalue()
        file_obj = io.BytesIO(file_content)
        wb = load_workbook(file_obj)
        sh = wb["Sheet1"]

        self.assertEqual(sh["A1"].value, "Financial year")
        self.assertEqual(sh["H2"].value, "Annually")
        self.assertEqual(sh["A3"].value, None)


class repetitive_API_Test(APITestCase):
    list_url = "/performance/api/v1/eqprs/"
    file_url = "/performance/performance-indicators.xlsx/"
    fixtures = ["test_repetitive_data.json"]

    def test_response_with_repetitive_data(self):
        api_response_payload = self.client.get(self.list_url).json()

        file_response = self.client.get(self.file_url, stream=True)
        file_content = file_response.getvalue()
        file_obj = io.BytesIO(file_content)
        wb = load_workbook(file_obj)

        sh = wb["Sheet1"]

        api_result_length = len(api_response_payload["results"]["items"])
        file_result_length = sh.max_row
        header_row_count = 1

        self.assertEqual(file_result_length - header_row_count, api_result_length)
