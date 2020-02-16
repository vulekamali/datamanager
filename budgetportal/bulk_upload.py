"""
Admin View for bulk uploading
"""

import logging
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from slugify import slugify

from .tasks import * 
from budgetportal.models import Department, Government, Sphere
from .datasets import (
    Category,
    Dataset,
    PackageDeletedException,
    PackageWithoutGroupException,
)
from django import forms
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render
from django_q.brokers import get_broker
from django_q.tasks import async_task

logger = logging.getLogger(__name__)

HEADINGS = [
    {"label": "government", "comment": None,},
    {"label": "group_id", "comment": None,},
    {"label": "department_name", "comment": None,},
    {
        "label": "dataset_name",
        "comment": 'This will be "sluggified" and must then be unique to this dataset in the entire system. For example, Gauteng Provincial Legislature EPRE for 2017-19 is prov-dept-gt-gauteng-provincial-legislature-2017-18',
    },
    {"label": "dataset_title", "comment": None,},
    {"label": "resource_name", "comment": None,},
    {
        "label": "resource_format",
        "comment": "You can usually make this the capitalised extension of the file. We use the combination of the format and resource title to ensure we do not duplicate resources. So we assume a file has only one resource with the same title and format combination.",
    },
    {"label": "resource_url", "comment": None,},
]


class FileForm(forms.Form):
    sphere_queryset = Sphere.objects.all()
    sphere = forms.ModelChoiceField(queryset=sphere_queryset, empty_label=None)
    file = forms.FileField(required=False)


def template_view(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resources"

    for idx, heading in enumerate(HEADINGS):
        column_letter = get_column_letter(idx + 1)
        ref = "%s1" % column_letter
        # Value
        ws[ref] = heading["label"]

        # Comment
        if heading["comment"]:
            ws[ref].comment = Comment(heading["comment"], "vulekamali")

        # Column width
        column = ws.column_dimensions[column_letter]
        column.width = 20

        # Style
        ws[ref].font = Font(bold=True)

    response = HttpResponse(
        content=save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=vulekamali-bulk-upload.xlsx"
    return response


def bulk_upload_view(request):
    form = None
    file = None
    valid = None
    preview = None

    if request.method == "POST":
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            valid = True
            sphere = form.cleaned_data["sphere"]
            if form.cleaned_data.get("file"):
                metadata_file = request.FILES["file"]
                preview = Preview(sphere, metadata_file)
            else:
                action_count = queue_actions(request.POST)
                if action_count:
                    messages.add_message(
                        request, messages.INFO, "Queued %d actions" % action_count
                    )
                else:
                    messages.add_message(
                        request,
                        messages.WARNING,
                        "No metadata file uploaded and no actions queued.",
                    )
            form = FileForm(initial={"sphere": sphere.pk})
        else:
            valid = False
    else:
        valid = True
        form = FileForm()

    return render(
        request,
        "admin/bulk_upload.html",
        {
            "valid": valid,
            "file": file,
            "form": form,
            "preview": preview,
            "queue_size": get_broker().queue_size(),
        },
    )


class Preview:
    def __init__(self, sphere, metadata_file):
        self.rows = []

        with BytesIO(metadata_file.read()) as bytes_io:
            wb = load_workbook(filename=bytes_io, read_only=True)
            ws = wb["Resources"]

            for row_idx, ws_row in enumerate(ws.rows):
                government = None
                department = None
                dataset = None

                if not ws_row[0].value:
                    continue

                if row_idx == 0:
                    heading_index = {}
                    for i, cell in enumerate(ws_row):
                        if cell.value:
                            heading_index[cell.value] = i
                else:
                    government_name = ws_row[heading_index["government"]].value
                    department_name = ws_row[heading_index[u"department_name"]].value
                    group_name = max_length_slugify(
                        ws_row[heading_index["group_id"]].value
                    )
                    dataset_name = max_length_slugify(
                        ws_row[heading_index["dataset_name"]].value
                    )
                    dataset_title = ws_row[heading_index["dataset_title"]].value
                    resource_name = ws_row[heading_index["resource_name"]].value
                    resource_format = ws_row[heading_index["resource_format"]].value
                    resource_url = ws_row[heading_index["resource_url"]].value

                    row = {}

                    government, government_preview = self.get_government_preview(
                        government_name, sphere
                    )
                    row["government"] = government_preview
                    department, department_preview = self.get_department_preview(
                        department_name, government
                    )
                    row["department"] = department_preview
                    dataset, dataset_preview = self.get_dataset_preview(
                        dataset_name, dataset_title, group_name, department
                    )
                    row["dataset"] = dataset_preview
                    group, group_preview = self.get_group_preview(group_name)
                    row["group"] = group_preview
                    resource, resource_preview = self.get_resource_preview(
                        dataset,
                        dataset_preview,
                        resource_name,
                        resource_format,
                        resource_url,
                    )
                    row["resource"] = resource_preview

                    self.rows.append(row)

    @classmethod
    def get_government_preview(cls, government_name, sphere):
        government = None
        government_preview = None

        government = Government.objects.filter(sphere=sphere, name=government_name)
        if government:
            government = government[0]
            government_preview = {
                "object": government,
                "name": government.name,
                "status": "success",
            }
        else:
            government_preview = {
                "name": government_name,
                "message": "Government not found for selected sphere",
                "status": "error",
            }
        return government, government_preview

    @classmethod
    def get_department_preview(cls, department_name, government):
        department = None
        department_preview = None

        if government:
            department = Department.objects.filter(
                government=government, name=department_name
            )
            if department:
                department = department[0]
                department_preview = {
                    "object": department,
                    "name": department.name,
                    "status": "success",
                }
            else:
                department_preview = {
                    "name": department_name,
                    "message": "Department not found for specified government",
                    "status": "error",
                }
        else:
            department_preview = {
                "name": department_name,
                "message": "Can't look up department without government",
                "status": "error",
            }
        return department, department_preview

    @classmethod
    def get_dataset_preview(cls, dataset_name, dataset_title, group_name, department):
        dataset = None
        dataset_preview = None
        try:
            if department:
                # Dataset for this department by group name
                dataset = department.get_dataset(group_name=group_name)
                if dataset:
                    dataset_preview = {
                        "object": dataset,
                        "name": dataset.slug,
                        "title": dataset.name,
                        "status": "success",
                        "message": "This dataset already exists",
                    }
                else:
                    try:
                        # Dataset by dataset name
                        dataset = Dataset.fetch(dataset_name)
                    except PackageDeletedException:
                        return (
                            None,
                            {
                                "name": dataset_name,
                                "title": dataset_title,
                                "status": "error",
                                "message": (
                                    "A dataset with this name exists but is deleted. "
                                    "First purge it or change its state to active or "
                                    "change its name"
                                ),
                            },
                        )
                    if dataset:
                        dataset_preview = {
                            "object": dataset,
                            "name": dataset.slug,
                            "title": dataset.name,
                            "new_title": dataset_title,
                            "status": "error",
                            "message": (
                                "Dataset by this name exists but it "
                                "is not configured correctly to be "
                                "identified as part of this dataset."
                            ),
                        }
                    else:
                        dataset_preview = {
                            "name": dataset_name,
                            "title": dataset_title,
                            "status": "info",
                            "message": "This dataset will be created.",
                            "action": "create",
                        }
            else:
                dataset_preview = {
                    "name": dataset_name,
                    "title": dataset_title,
                    "status": "error",
                    "message": (
                        "Department not found. We can't "
                        "upload the dataset until we can "
                        "associate it with an existing department."
                    ),
                }
            return dataset, dataset_preview
        except PackageWithoutGroupException:
            return (
                None,
                {
                    "name": dataset_name,
                    "title": dataset_title,
                    "status": "error",
                    "message": (
                        "Dataset exists but must have the correct group " "assigned."
                    ),
                },
            )

    @classmethod
    def get_group_preview(cls, group_name):
        category = None
        group_preview = None
        category = Category.get_by_slug(group_name)
        if category:
            group_preview = {
                "name": category.slug,
                "object": category,
                "status": "success",
            }
        else:
            group_preview = {
                "name": group_name,
                "status": "error",
                "message": (
                    "Group not found. Ensure that group exists and "
                    "correct name is used in your metadata."
                ),
            }
        return category, group_preview

    @classmethod
    def get_resource_preview(cls, dataset, dataset_preview, name, format, url):
        resource = None
        resource_preview = None
        if dataset:
            resource = dataset.get_resource(format, name)
            if resource:
                resource_preview = {
                    "name": name,
                    "format": format,
                    "url": url,
                    "status": "success",
                    "message": "This resource already exists.",
                }
            else:
                resource_preview = {
                    "name": name,
                    "format": format,
                    "url": url,
                    "status": "info",
                    "message": "This resource will be created.",
                    "action": "create",
                }
        else:
            if dataset_preview.get("action", None) == "create":
                resource_preview = {
                    "name": name,
                    "format": format,
                    "url": url,
                    "status": "info",
                    "message": "This resource will be created after the dataset is created.",
                    "action": "create",
                }
            else:
                resource_preview = {
                    "name": name,
                    "format": format,
                    "url": url,
                    "status": "error",
                    "message": "Can't create resource without a matching dataset.",
                }
        return resource, resource_preview


def queue_actions(post_data):
    action_count = 0
    # Using kwargs to django-q so it shows field names in Admin
    for preview_idx, action in enumerate(post_data.getlist("dataset_action[]")):
        if action == "create":
            async_task(
                create_dataset,
                **{
                    "department_id": post_data.getlist("department_id[]")[preview_idx],
                    "name": post_data.getlist("dataset_name[]")[preview_idx],
                    "title": post_data.getlist("dataset_title[]")[preview_idx],
                    "group_name": post_data.getlist("group_name[]")[preview_idx],
                }
            )
            action_count += 1
    for preview_idx, action in enumerate(post_data.getlist("resource_action[]")):
        if action == "create":
            async_task(
                create_resource,
                **{
                    "department_id": post_data.getlist("department_id[]")[preview_idx],
                    "dataset_name": post_data.getlist("dataset_name[]")[preview_idx],
                    "group_name": post_data.getlist("group_name[]")[preview_idx],
                    "name": post_data.getlist("resource_name[]")[preview_idx],
                    "format": post_data.getlist("resource_format[]")[preview_idx],
                    "url": post_data.getlist("resource_url[]")[preview_idx],
                }
            )
            action_count += 1
    return action_count


def max_length_slugify(value):
    return slugify(value, max_length=85, word_boundary=True)
